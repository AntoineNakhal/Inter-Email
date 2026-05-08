"""Excel (.xlsx) text extraction via openpyxl.

Each sheet becomes a labelled section. Cells are joined with " | " so the
embedding model sees something resembling Markdown tables (which it tokenizes
sensibly) rather than CSV (which it mostly treats as noise).

We use `data_only=True` so cached formula results are read instead of the
formulas themselves — no point embedding `=SUM(A1:A12)`.
"""

from __future__ import annotations

import io

from backend.knowledge.extractors.base import BaseExtractor, ExtractionError


class XlsxExtractor(BaseExtractor):
    file_type = "xlsx"

    def extract(self, content: bytes, *, filename: str = "") -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ExtractionError(
                "openpyxl is not installed — run `pip install openpyxl`."
            ) from exc

        try:
            workbook = load_workbook(
                filename=io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
            sections: list[str] = []
            for sheet in workbook.worksheets:
                rows_text: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [
                        str(cell).strip() if cell is not None else ""
                        for cell in row
                    ]
                    if any(cells):
                        rows_text.append(" | ".join(cells))
                if rows_text:
                    sections.append(
                        f"Sheet: {sheet.title}\n" + "\n".join(rows_text)
                    )
        except Exception as exc:
            raise ExtractionError(
                f"Failed to extract text from Excel '{filename}': {exc}"
            ) from exc

        return "\n\n".join(sections).strip()
